import csv
import math
import openpyxl

#modules = ["DLMA-1010","DLMB-1070"]
modules = module_names

MagnetFiducials = dict()

FiducialFileName = FiducialFile
# If you want to override the FiducialFileName, define it here
# example
# FiducialFileName="SourceFiles/APS_U_MagnetFiducials_All.xlsx"

# ModuleDataDirectory = "./ModuleDataDirectory"

workbook = openpyxl.load_workbook(FiducialFileName)
sheet_list = workbook.sheetnames
for sheet_name in sheet_list:
  if sheet_name == 'Legend':
    continue
  sheet = workbook[sheet_name]
  for row in sheet.iter_rows():
    values = []
    for cell in row:
      values.append(cell.value);
    if (values[0] == None) or (values[0] == 'Point ID'):
      continue
    if type(values[0]) != str :
      continue
#    print(type(values[0]))

    PointName = values[0].split("_")
    if (PointName[0] not in sheet_name) :
      continue
#    print(PointName)
    MagnetKey = PointName[0]
    MagnetID = PointName[1]
    FiducialID = PointName[2]

    X = float(values[1])
    Y = float(values[2])
    Z = float(values[3])
    
    if MagnetKey not in MagnetFiducials:
      MagnetFiducials[MagnetKey] = dict()

    if MagnetID not in MagnetFiducials[MagnetKey]:
      MagnetFiducials[MagnetKey][MagnetID] = dict()
    MagnetFiducials[MagnetKey][MagnetID][FiducialID] = [X,Y,Z]

#print(MagnetFiducials)

for module in modules:
  MagnetInfo = []

  InputFile = os.path.join(ModuleDataDirectory, module, module + ".csv")
  print()
  print("Reading InputFile: " ,InputFile)
    
  #OutDataDir + "/" + module + ".csv"
  with open(InputFile, 'r',) as infile:
    reader = csv.reader(infile, delimiter = ',')
    for row in reader:
        MagnetInfo.append(row)

#  print(TransformTable)
  module_type = module.split('-')[0]

  for magnet in MagnetInfo:
    OriginalCoordinates = []
    m_prefix = magnet[2].split(":")[0]
    m_type = magnet[2].split(":")[1]
    if 'FC' in m_type:
      continue
    if 'DLM' in m_type:
      continue
    if 'BUSBAR' in m_type:
      continue
    if 'SUPP' in m_type:
      continue
#    if 'M' in m_type:
#      continue
    
    m_id = magnet[4][-3:]

    m_type2 = m_type
    if m_type == 'S3':
      m_type2 = 'S1'


    #output_name = OutDataDir + "/" + m_type + "_" + m_id + ".csv"
    output_name = os.path.join(ModuleDataDirectory, module,"L_" + m_type + "_" + m_id + ".csv")
    print("Write to OutputFile: " , output_name)
#    print (output_name)

    for point in MagnetFiducials[m_type2][m_id]:
      X = MagnetFiducials[m_type2][m_id][point][0]
      Y = MagnetFiducials[m_type2][m_id][point][1]
      Z = MagnetFiducials[m_type2][m_id][point][2]

      out_row = []
      #LaticePosition = module[3]+m_type
      LaticePosition = m_prefix+m_type
      ModuleName = ""
      if 'DLM' in module:
        ModuleName = module.replace("LM","").replace("-","")
      elif 'FODO' in module:
        ModuleName = module.replace("O","").replace("-","")
      elif 'QMQ' in module:
        ModuleName = module.replace("MQ","").replace("-","")
      OutModuleName = ModuleName[0:2] + ModuleName[3:5]
      FiducialName = OutModuleName + "_"+LaticePosition+"_"+point

#out_row += [point,X,Y,Z]
      out_row += [FiducialName,X,Y,Z]

      OriginalCoordinates.append(out_row)

#  print(OriginalCoordinates)
    with open(output_name, 'w', newline='') as outfile:
      writer = csv.writer(outfile)
#      writer.writerow(["Tag", "Serial Number", "Machine Name" ,"Catalog Item" , "QRID" , "Transform ID" ,"dZ [m]" , "dX [m]" , "dY [m]","dTheta [rad]" ,"Point ID" , "X [m]" , "Y [m]",	"Z [m]" , "Ideal X [m]" , "Ideal Y [m]" , "Ideal Z [m]"])
      for row in OriginalCoordinates:
        writer.writerow(row);

    
