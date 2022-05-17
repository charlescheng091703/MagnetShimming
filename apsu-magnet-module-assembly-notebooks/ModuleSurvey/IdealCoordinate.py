import csv
import math
import openpyxl
from datetime import date
import os

modules = module_names
# If you want to override the modules, define it here
# example
#modules = ["DLMA-1010","DLMB-1070"]

MagnetFiducials = dict()
FiducialFileName = FiducialFile
LatticeFileName = LatticeFile
# If you want to override the FiducialFileName and LatticeFileName, define it here
# example
#FiducialFileName="SourceFiles/APS_U_MagnetFiducials_All.xlsx"
#LatticeFileName = "SourceFiles/Lattice_June2021revision.xlsx"

ModuleDataDirectory = "./ModuleDataDirectory"

workbook = openpyxl.load_workbook(FiducialFileName)

sheet_list = workbook.sheetnames
for sheet_name in sheet_list:
  if sheet_name == 'Legend':
    continue
  sheet = workbook[sheet_name]
  for row in sheet.iter_rows():
    values = []
    for cell in row:
      values.append(cell.value);
    if (values[0] == None) or (values[0] == 'Point ID'):
      continue
    if type(values[0]) != str :
      continue
#    print(type(values[0]))

    PointName = values[0].split("_")
    if (PointName[0] not in sheet_name) :
      continue
#    print(PointName)
    MagnetKey = PointName[0]
    MagnetID = PointName[1]
    FiducialID = PointName[2]

    X = float(values[1])
    Y = float(values[2])
    Z = float(values[3])
    
    if MagnetKey not in MagnetFiducials:
      MagnetFiducials[MagnetKey] = dict()

    if MagnetID not in MagnetFiducials[MagnetKey]:
      MagnetFiducials[MagnetKey][MagnetID] = dict()
    MagnetFiducials[MagnetKey][MagnetID][FiducialID] = [X,Y,Z]

#print(MagnetFiducials)

for module in modules:
  IdealCoordinates = []
  MagnetInfo = []
  TransformationInfo = []

  InputFile = os.path.join(ModuleDataDirectory, module, module +".csv")
  print("Processing File: ", InputFile)
  with open(InputFile, 'r',) as infile:
    reader = csv.reader(infile, delimiter = ',')
    for row in reader:
        MagnetInfo.append(row)

  TransformTable = dict()
  TransformTable['DLMA'] = dict()
  TransformTable['DLMB'] = dict()
  TransformTable['QMQA'] = dict()
  TransformTable['QMQB'] = dict()
  TransformTable['FODO'] = dict()

  lattice_workbook = openpyxl.load_workbook(LatticeFileName)
  lattice_sheet = lattice_workbook["A107-000021"]
  
  for row in lattice_sheet.iter_rows():
    values = []
    for cell in row:
      values.append(cell.value);
    if (values[0] == None) or (values[0] == 'MODULE'):
      continue
    if type(values[0]) != str :
      continue
    #print(type(values[0]))
    #print(values[0])
    #if not('A:' in values[0]) and not('B:' in values[0]) :
    #  continue
    mod = values[0]
    substr = values[1][2:].split()
    key = substr[0]
    ref = substr[1]

    z = float(values[2])/1000.0
    x = - float(values[3])/1000.0
    theta = -float(values[4])*math.pi/180.0
    
    if not(mod in TransformTable):
      print("Module ",mod," is not recognoized.")
    
    TransformTable[mod][key] = [ref,z,x,theta]
    #if 'A:' in values[0]:
    #  TransformTable['DLMA'][key] = [ref,z,x,theta]
    #if 'B:' in values[0]:
    #  TransformTable['DLMB'][key] = [ref,z,x,theta]

  #print(TransformTable)
  module_type = module.split('-')[0]
    
  ModuleName = ""
  if 'DLM' in module:
    ModuleName = module.replace("LM","").replace("-","")
  elif 'FODO' in module:
    ModuleName = module.replace("O","").replace("-","")
  elif 'QMQ' in module:
    ModuleName = module.replace("MQ","").replace("-","")
  OutModuleName = ModuleName[0:2] + ModuleName[3:5]
  output_name = os.path.join(ModuleDataDirectory, module, OutModuleName + "_IdealCoordinates.csv")
  print("Write to OutputFile: ", output_name, "\n")

  for magnet in MagnetInfo:
    IndividualIdealCoordinates = []
    
    m_prefix = magnet[2].split(":")[0]
    m_type = magnet[2].split(":")[1]
    if 'FC' in m_type:
#      out_row = magnet
#      IdealCoordinates.append(out_row)
      continue
    if 'DLM' in m_type:
      continue
    if 'SUPP' in m_type:
      continue
    if 'BUSBAR' in m_type:
      continue
    
    transform_info = TransformTable[module_type][m_type]
    dz = transform_info[1]
    dx = transform_info[2]
    dtheta = transform_info[3]

    m_id = magnet[4][-3:]

    m_type2 = m_type
    if m_type == 'S3':
      m_type2 = 'S1'
    for point in MagnetFiducials[m_type2][m_id]:
      X = MagnetFiducials[m_type2][m_id][point][0]
      Y = MagnetFiducials[m_type2][m_id][point][1]
      Z = MagnetFiducials[m_type2][m_id][point][2]

      Z_ideal = Z*math.cos(dtheta) - X*math.sin(dtheta) + dz
      X_ideal = Z*math.sin(dtheta) + X*math.cos(dtheta) + dx
      Y_ideal = Y

      point_id = m_type2+"_"+m_id+"_"+point
      #LaticePosition = module[3]+m_type
      LaticePosition = m_prefix+m_type
      FiducialName = OutModuleName+"_"+LaticePosition+"_"+point
#      print (FiducialName)

#      print(point_id,X,Y,Z)

      out_row = []
#      out_row += magnet
#      out_row += [module[3]+":"+m_type+"_"+transform_info[0],dz,dx,0,dtheta,point_id,X,Y,Z,X_ideal,Y_ideal,Z_ideal]
      out_row += [FiducialName]
      out_row += [X_ideal,Y_ideal,Z_ideal]

      IdealCoordinates.append(out_row)
    
      IndividualIdealCoordinates.append(out_row)
        
    IndividualOutputName = os.path.join(ModuleDataDirectory, module,
    "Z_" + m_type + "_" + m_id + ".csv")
    with open(IndividualOutputName, 'w', newline='') as outfile:
      writer = csv.writer(outfile)
      for row in IndividualIdealCoordinates:
        writer.writerow(row);

#  print(IdealCoordinates)
  with open(output_name, 'w', newline='') as outfile:
    writer = csv.writer(outfile)
#    writer.writerow(["Tag", "Serial Number", "Machine Name" ,"Catalog Item" , "QRID" , "Transform ID" ,"dZ [m]"
#    , "dX [m]" , "dY [m]","dTheta [rad]" ,"Point ID" , "X [m]" , "Y [m]",	"Z [m]" , "Ideal X [m]" , "Ideal Y [m]" , "Ideal Z [m]"])
    for row in IdealCoordinates:
      writer.writerow(row);

    
