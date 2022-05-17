# %%
#!/usr/bin/env python

import numpy as np
import scipy.stats as scistats
import pandas as pd

# Note: multiprocess AND openpyxl also need to be imported.
import time, os, math, itertools, csv, glob, re
from multiprocess import Pool
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import ipywidgets as widgets
from IPython.display import display, clear_output
from ipyfilechooser import FileChooser
from colorama import Fore, Style
from pathlib import Path
from datetime import datetime
import os
import csv
import requests
import copy
# Try and bring in rich if we have it.
try:
    from rich import print
except:
    pass

var_dict = {}
VERSION = "20220221.0"


def moduleToCDBInfoFilename(module, pathdir="CDBInfo"):
    cdbInfoDir = Path(pathdir)
    filename = module + ".xlsx"
    cdbInfoFilename = str(cdbInfoDir / filename)
    # This is a hack to deal with an anomoaly that Laura is Seeing
    # Replace any \\ with a \
    cdbInfoFilename.replace("\\\\", "\\")
    return cdbInfoFilename


def listCDBInfoFiles(pathdir="CDBInfo"):
    cdbInfoDir = Path(pathdir)
    xlsxdir = cdbInfoDir / "*.xlsx"
    #     files = glob.glob(str(xlsxdir))
    filelist = [file.replace("\\\\", "\\") for file in glob.glob(str(xlsxdir))]
    return filelist


try:
    # NOTE: Raise the exception below to prevent accessing the CDB
    # if you have CDB Libraries loaded
    #    raise(ImportError)
    from CdbApiFactory import CdbApiFactory

    # We only consider Q and S Magnets in the calculation

    dlma_module_distance = {}
    dlma_module_distance["A:Q1"] = 3033.09420075
    #    dlma_module_distance['A:FC1'] = 3414.17802624
    dlma_module_distance["A:Q2"] = 3675.67802624
    #    dlma_module_distance['A:M1'] = 4652.61256868
    dlma_module_distance["A:Q3"] = 6301.30749401
    dlma_module_distance["A:S1"] = 6647.11440155
    dlma_module_distance["A:S2"] = 7519.80296993
    dlma_module_distance["A:Q4"] = 7150.03849285
    dlma_module_distance["A:Q5"] = 7865.27388654
    #    dlma_module_distance['A:FC2'] = 8098.23134352
    dlma_module_distance["A:S3"] = 8379.82901440

    qmqa_module_distance = {}
    qmqa_module_distance["A:Q6"] = 8755.08390852
    qmqa_module_distance['A:M2'] = 10046.84917829
    qmqa_module_distance["A:Q7"] = 11162.59772205

    fodo_module_distance = {}
    fodo_module_distance["A:M3"] = 11999.33633188
    fodo_module_distance["A:Q8"] = 12845.06277413
    fodo_module_distance["A:M4"] = 13786.49242732
    fodo_module_distance["B:Q8"] = 14726.46836163
    fodo_module_distance["B:M3"] = 15571.60143966

    qmqb_module_distance = {}
    qmqb_module_distance["B:Q7"] = 16404.45760145
    qmqb_module_distance['B:M2'] = 17515.02909743
    qmqb_module_distance["B:Q6"] = 18796.08655755

    dlmb_module_distance = {}
    dlmb_module_distance["B:S3"] = 19168.23085714
    #    dlmb_module_distance['B:FC2'] = 19447.49428486
    dlmb_module_distance["B:Q5"] = 19678.52069124
    dlmb_module_distance["B:S2"] = 20021.19049956
    dlmb_module_distance["B:Q4"] = 20387.95689404
    dlmb_module_distance["B:S1"] = 20886.93762729
    dlmb_module_distance["B:Q3"] = 21230.03311082
    #    dlmb_module_distance['B:M1'] = 22865.80084786
    dlmb_module_distance["B:Q2"] = 23830.70770496
    #    dlmb_module_distance['B:FC1'] = 24088.98820603
    dlmb_module_distance["B:Q1"] = 24465.38025725

    CDBItemID = {}
    CDBItemID["DLMA"] = 110353
    CDBItemID["DLMB"] = 110354
    CDBItemID["FODO"] = 110371
    CDBItemID['QMQA'] = 110369
    CDBItemID['QMQB'] = 110370

    ModuleDistances = {}
    ModuleDistances["DLMA"] = dlma_module_distance
    ModuleDistances["DLMB"] = dlmb_module_distance
    ModuleDistances["FODO"] = fodo_module_distance
    ModuleDistances["QMQA"] = qmqa_module_distance
    ModuleDistances["QMQB"] = qmqb_module_distance

    def numberOfAssignedQrCodes(filename_module):
        cdbitem = MagnetModules[filename_module].id
        item_hierarchyOBJ = itemApi.get_item_hierarchy_by_id(cdbitem)
        # We are going to create a list of dicts for all if the items that have
        # a QR Code in the assembly
        qrCodedItemList = []
        for item_hierarchy in item_hierarchyOBJ.child_items:
            currentQrCodedItem = {}
            if item_hierarchy != None:
                currentQrCodedItem["Magnet"] = item_hierarchy.derived_element_name
                if item_hierarchy.item != None:
                    if item_hierarchy.item.qr_id:
                        currentQrCodedItem["QR Code"] = item_hierarchy.item.qr_id
                        qrCodedItemList.append(currentQrCodedItem)
        return len(qrCodedItemList)

    #    print("Connection test to cdb server")
    #    response = requests.get("https://cdb.aps.anl.gov/")
    #    print(response)

    apiFactory = CdbApiFactory("https://cdb.aps.anl.gov/cdb")
    itemApi = apiFactory.getItemApi()

    # Preload the allowed Magnet Modules with the list from CDB
    MagnetModules = {}
    MagnetModuleNames = []
    MagnetModuleDistances = {}
    # Let's store the magnets by name.  This isn't the best but it will work for now
    # Since we will be pulling magnet information out of the CDB and putting it into
    # data files, we will use the filename as the key to store the CDB Info.
    for magnet_module in CDBItemID.keys():
        for inv_item in itemApi.get_items_derived_from_item_by_item_id(
            CDBItemID[magnet_module]
        ):
            modulekey = moduleToCDBInfoFilename(inv_item.name)
            MagnetModules[modulekey] = inv_item
            MagnetModuleDistances[modulekey] = ModuleDistances[magnet_module]
            if numberOfAssignedQrCodes(modulekey) > 1:
                MagnetModuleNames.append(inv_item.name)

    MagnetModuleNames = sorted(MagnetModuleNames)
    USE_CDBAPI = True
    print("CDBApi Is Loaded")


except:
    USE_CDBAPI = False
    MagnetModuleNames = []
    filelist = listCDBInfoFiles()
    for name in filelist:
        prefix = os.path.basename(name)
        MagnetModuleNames.append(prefix.replace(".xlsx", ""))
    MagnetModuleNames = sorted(MagnetModuleNames)
    print("CDBApi Not Loaded")


# %%
def init_worker(permutations, X_np, Z_np, MagShim_np, stderr_weight):

    """
    init_worker assigns its arguments to keys in var_dict

    :var_dict: a global dictionary storing variables passed from the initializer 

    """

    var_dict["permutations"] = permutations
    var_dict["X_np"] = X_np
    var_dict["Z_np"] = Z_np
    var_dict["MagShim_np"] = MagShim_np
    var_dict["stderr_weight"] = stderr_weight


# %%
def lin_regress(index_shim_numbers):

    """
    lin_regress performs a linear regression on the iterated set of coordinate points and outputs the standard error 

    """
    shim_numbers = var_dict["permutations"][index_shim_numbers]
    X_tofit = var_dict["X_np"] + np.multiply(shim_numbers, var_dict["MagShim_np"])
    regression_result = scistats.linregress(var_dict["Z_np"], X_tofit)
    stderr = float(regression_result.stderr)
    weighted_score = var_dict["stderr_weight"] * stderr / 5e-6 + (
        1 - var_dict["stderr_weight"]
    ) * np.sum(shim_numbers) / (len(shim_numbers) * 5)

    return [stderr, weighted_score]

def lin_regress_zero_slope(index_shim_numbers):

    """
    lin_regress performs a linear regression on the iterated set of coordinate points and outputs the standard error 

    """
    shim_numbers = var_dict["permutations"][index_shim_numbers]
    X_tofit = var_dict["X_np"] + np.multiply(shim_numbers, var_dict["MagShim_np"])
    regression_result = scistats.linregress(var_dict["Z_np"], X_tofit)
    slope = abs(float(regression_result.slope))
    weighted_score = var_dict["stderr_weight"] * slope / 5e-6 + (
        1 - var_dict["stderr_weight"]
    ) * np.sum(shim_numbers) / (len(shim_numbers) * 5)

    return [slope, weighted_score]

# %%
def zero_in_set(i):

    """
    zero_in_set returns true if the list contains the value zero and false if it doesn't 

    """

    return 0 in i


# %%
def neg_in_set(i, NegShim):

    """
    neg_in_set returns true if the list i contains at least one value in NegShim and false if it doesn't 

    """

    return any(i == NegShim)


# %%
def plot(x, y1, y2, axis, title):

    """
    plot produces a scatter graph of the relationships y1 vs. x and y2 vs. x; axis labels, legend,
    and title can be modified 

    """

    plt.figure()
    plt.scatter(x, y1, marker="+", s=30, label="Raw d" + axis)
    z1 = np.polyfit(x, y1, 1)
    p1 = np.poly1d(z1)
    plt.plot(x, p1(x))
    plt.scatter(x, y2, marker="o", s=15, label="Shimmed d" + axis)
    z2 = np.polyfit(x, y2, 1)
    p2 = np.poly1d(z2)
    plt.plot(x, p2(x))
    plt.xlabel("Z")
    plt.ylabel(axis + " Error")
    plt.legend(bbox_to_anchor=(1.05, 1.0), loc="upper left")
    plt.title(title)
    plt.show()


# %%
def index_of_FC_and_M(crd_name):

    """
    index_of_FC_and_M returns the indices of elements in the list crd_name containing "FC" or "M"

    """

    return [crd_name.index(i) for i in crd_name if "FC" in i or "M" in i]


def createMagnetFileFromCDB(filename_module, outwidget):
    try:
        cdbitem = MagnetModules[filename_module].id
        item_hierarchyOBJ = itemApi.get_item_hierarchy_by_id(cdbitem)
        # We are going to create a list of dicts for all if the items that have
        # a QR Code in the assembly
        with outwidget:
            qrCodedItemList = []
            for item_hierarchy in item_hierarchyOBJ.child_items:
                currentQrCodedItem = {}
                if item_hierarchy != None:
                    currentQrCodedItem["Magnet"] = item_hierarchy.derived_element_name
                    if item_hierarchy.item != None:
                        if item_hierarchy.item.qr_id:
                            currentQrCodedItem["QR Code"] = item_hierarchy.item.qr_id
                            qrCodedItemList.append(currentQrCodedItem)
            qrDataFrame = pd.DataFrame(qrCodedItemList)
            # Get List of magnet ids tha we care about
            magnetstofit = qrDataFrame[
                qrDataFrame["Magnet"].isin(
                    MagnetModuleDistances[filename_module].keys()
                )
            ].copy(deep=True)
            magnetstofit["Z mm"] = magnetstofit["Magnet"].map(
                MagnetModuleDistances[filename_module]
            )
            magnetstofit = magnetstofit.sort_values(by=["Z mm"])
            magnetstofit.to_excel(filename_module)
    except:
        with outwidget:
            print("DATA FILE NOT CREATED PROPERLY.  DO NOT RUN CALCULATION")


def refreshMagnetConfigurationDataFromCDBIfPossible(filename_module, outwidget):
    if USE_CDBAPI:
        with outwidget:
            print("Pulling Magnet Configuration Data From CDB into CDBInfo")
        createMagnetFileFromCDB(filename_module, outwidget)
    else:
        with outwidget:
            print("Not Pulling Magnet Configuration Data From CDB into CDBInfo")


# %%
def extract_shim_data_it1(module_name, filename_module, filename_fiducials, outwidget):

    """ 
    extract_shim_data reads the QR codes and Z coordinates from filename_module, matches the QR code with its corresponding point ID in filename_fiducials, and extracts the x and y offsets of each point. Returns dictionary with Z coordinates, offsets in the x axis, offsets in the y axis, QR code numbers, and coordinate names. 
    
    """

    # If the CDB API is Active, we are going to construct the file on the fly. The data will be read out of the CDB and tech file
    # in filename_module will be overwritten with the data.

    var_dict = {}

    filename_lattice = filename_module
    try:
        wb_lattice = load_workbook(filename_lattice, data_only=True)
    except:
        with outwidget:
            print("Error: Can't read ", filename_lattice, " as an Excel File")
        return None
    ws_mod = wb_lattice.active

    col_names_mod = [c.value for c in next(ws_mod.iter_rows(min_row=1, max_row=1))]
    qr_code_mod = [
        row_cells[col_names_mod.index("QR Code")].value
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]
    qr_code_mod = [i for i in qr_code_mod if i]
    crd_name = [
        row_cells[col_names_mod.index("Magnet")].value
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]
    crd_name = [i for i in crd_name if i]
    # index_FCandM = index_of_FC_and_M(crd_name)
    Z = [
        row_cells[col_names_mod.index("Z mm")].value
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]
    Z = [i for i in Z if i]
    var_dict["QR_Code_Num"] = qr_code_mod
    var_dict["Magnet"] = crd_name

    filename_fiducials = filename_fiducials
    try:
        wb_fiducials = load_workbook(filename_fiducials, data_only=True)
        with outwidget:
            print("Fiducials File Read")
    except:
        with outwidget:
            print("Error: Can't read ", filename_fiducials, " as an Excel File")
        return None
    ws_CDB_XRef = wb_fiducials["CDB Xref"]
    col_names_CDB_XRef = [
        c.value for c in next(ws_CDB_XRef.iter_rows(min_row=1, max_row=1))
    ]
    qr_code_searchlist = [
        row_cells[col_names_CDB_XRef.index("CDB QR Code")].value
        for row_cells in ws_CDB_XRef.iter_rows(min_row=2, max_row=ws_CDB_XRef.max_row)
    ]
    try:
        qr_indices_mod = []
        for i in qr_code_mod:
            qr_code_current_search = i
            qr_indices_mod.append(qr_code_searchlist.index(i))
    except ValueError:
        with outwidget:
            print(
                Fore.RED
                + "Error: QR code number not found on fiducial file cross reference sheet."
                + Style.RESET_ALL
            )
            print(
                Fore.RED
                + "QR Code Not Found is: "
                + str(qr_code_current_search)
                + Style.RESET_ALL
            )
            return None
    with outwidget:
        print("All Files Read Correctly")
    surv_dsgntr_mod = [
        ws_CDB_XRef.cell(i + 2, col_names_CDB_XRef.index("CDB QR Code")).value
        for i in qr_indices_mod
    ]

    dX = []
    dY = []

    # Assume we can parse all of the data.
    survey_data_parsed = True
    for i in surv_dsgntr_mod:
        if i[0:2] == "S1" or i[0:2] == "S3":
            ws_pt = wb_fiducials["S1_S3"]
        else:
            ws_pt = wb_fiducials[i[0:2]]

        col_names_pt = [c.value for c in next(ws_pt.iter_rows(min_row=1, max_row=1))]
        ptid_searchlist = [row_cells[col_names_pt.index("Point ID")].value for row_cells in ws_pt.iter_rows(min_row=2, max_row=ws_pt.max_row)]
        
        surveyid_i = i + "_0"
        if i[0:2] != 'M1' and i[0:2] != 'M2' and surveyid_i in ptid_searchlist:
            dX_pt = ws_pt.cell(ptid_searchlist.index(i + "_0") + 2, col_names_pt.index("dX [mm]") + 1).value
        elif surveyid_i in ptid_searchlist:
            dX_pt = (ws_pt.cell(ptid_searchlist.index(i + "_0") + 2, col_names_pt.index("dX_DS [mm]") + 1).value + ws_pt.cell(ptid_searchlist.index(i + "_0") + 2, col_names_pt.index("dX_US [mm]") + 1).value)/2.0
        else:
            with outwidget:
                print(
                    Fore.RED
                    + "Error: Survey Point "
                    + surveyid_i
                    + " is not in the fiducials file."
                    + Style.RESET_ALL
                )
            survey_data_parsed = False
            continue

        if dX_pt == None:
            dX_pt = float("inf")
        dY_pt = ws_pt.cell(
            ptid_searchlist.index(i + "_0") + 2, col_names_pt.index("dY [mm]") + 1
        ).value
        if dY_pt == None:
            dY_pt = float("inf")

        dX.append(dX_pt)
        dY.append(dY_pt)

    Znp = (1.0e-3) * np.array(Z)
    dXnp = (1.0e-3) * np.array(dX)
    dYnp = (1.0e-3) * np.array(dY)
    var_dict["Z"] = Znp
    var_dict["X"] = dXnp
    var_dict["Y"] = dYnp
    var_dict["Shim Pack/Change Step"] = [0 for i in range(len(var_dict["Z"]))]

    if survey_data_parsed:
        return var_dict
    else:
        return None


def parse_survey_module_name(surveykey):
    """ The Survey Group has a different designation for the 
    Module Names in their files.   This will parese the key
    and return a list where the first part is the module
    and the second is the magnet """
    # Establsh the keys for the moduleLookup
    moduleLookup = {}
    moduleLookup["DA"] = "DLMA"
    moduleLookup["DB"] = "DLMB"
    moduleLookup["QA"] = "QMQA"
    moduleLookup["QB"] = "QMQB"
    moduleLookup["FD"] = "FODO"

    tokens = surveykey.split("_")
    # Get the magnet.  Keep the module designation
    magnet = tokens[1]
    # Get the module type
    moduletype = moduleLookup[tokens[0][0:2]]
    modulenumber = int(tokens[0][2:]) * 10 + 1000
    modulestr = moduletype + "-" + str(modulenumber)
    return (modulestr, magnet)


def extract_shim_data_it2(module_name, filename_module, filename_fiducials, outwidget):

    """ 
    extract_shim_data_it2 takes the module type from the filename_module variable, reads in the as measured data from the filename_fiducials file
    given by Bill Jansma and returns the data from that. 
    
    """

    # If the CDB API is Active, we are going to construct the file on the fly. The data will be 
    # read out of the CDB and tech file
    # in filename_module will be overwritten with the data.

    var_dict = {}

    # Read in the Module Identification from the CDBInfo file
    # This data is not used to look up information but it will be used to get the names of 
    # the magnet array
    filename_lattice = filename_module
    try:
        wb_lattice = load_workbook(filename_lattice, data_only=True)
    except:
        with outwidget:
            print("Error: Can't read ", filename_lattice, " as an Excel File")
        return None

    ws_mod = wb_lattice.active
    col_names_mod = [c.value for c in next(ws_mod.iter_rows(min_row=1, max_row=1))]
    qr_code_mod = [
        row_cells[col_names_mod.index("QR Code Number")].value
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]
    qr_code_mod = [i for i in qr_code_mod if i]
    magnet_col_name = module_name + " Magnets"
    crd_name = [
        row_cells[col_names_mod.index(magnet_col_name)].value
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]
    crd_name = [i for i in crd_name if i]
    # index_FCandM = index_of_FC_and_M(crd_name)
    Z = [
        row_cells[col_names_mod.index("Z")].value
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]
    Z = [i for i in Z if i]

    adjustment_steps = [
        int(row_cells[col_names_mod.index("Shim Pack/Change Step")].value) + 1
        for row_cells in ws_mod.iter_rows(min_row=2, max_row=ws_mod.max_row)
    ]


    # So, at this point,   We have the qr_code_mod and crd_name loaded with the Module data and the Z files from
    # the CDBInfo File that was generated.  Note, these are currently lists.
    # Our approach will be to read the fiducials filename and attempt to correlate the keys in the
    # file with the keys that are in crd_name.   If we have a match, we will add the dx and dy
    # to the data.    If we don't, we will remove the entry from the var_dict

    # Look for common naming keys between the CDB and modula assembly survey name spaces and return the intersection

    # We need to remap the cdbIdList form the name, by deleting the :
    # in the Assembly name.
    cdbIdList = [val.replace(":","") for val in crd_name]
    moduleIdList = []
    magnetIdList = []
    moduledX = []
    moduledY = []
    moduledZ = []


    # Attempt to read the csv file.  Any error abort
    try:
        with open(filename_fiducials) as csv_file:
            csv_reader = csv.reader(csv_file, delimiter=",")
            for row in csv_reader:
                if len(row) == 0:
                    break
                # Reduce row [0] to the type of magnet
                moduleId, magnetId = parse_survey_module_name(row[0])
                moduleIdList.append(moduleId)
                magnetIdList.append(magnetId)
                # Note:  We have to flip the X Coordinate to Match email specifications from Jeremy
                moduledX.append(-float(row[1]))
                moduledY.append(float(row[2]))
                moduledZ.append(float(row[3]))
            allowedelements = [
                [i, j]
                for i, cdb in enumerate(cdbIdList)
                for j, mod in enumerate(magnetIdList)
                if cdb == mod
            ]
    except:
        with outwidget:
            print("Error reading in the Survey Data File")
        return None

    # Error Check.  See if all of the ModuleIds in the Survey File match the Survey Name

    modulenamecheck = [module_name == moduleId for moduleId in moduleIdList]
    if all(modulenamecheck):
        with outwidget:
            print("Survey Data matches selected Module Name: ", module_name)
    else:
        with outwidget:
            print("Survey Data does not match Module Name: ", module_name)
            print("Survey Data Keys designate Modules: ", moduleIdList)
        return

    var_dict["QR_Code_Num"] = [qr_code_mod[itsok[0]] for itsok in allowedelements]
    var_dict["Magnet"] = [crd_name[itsok[0]] for itsok in allowedelements]
    dX = [moduledX[itsok[1]] for itsok in allowedelements]
    dY = [moduledY[itsok[1]] for itsok in allowedelements]
    Z = [Z[itsok[0]] for itsok in allowedelements]
    dZ = [moduledZ[itsok[1]] for itsok in allowedelements]

    ###
    dXnp = np.array(dX)
    dYnp = np.array(dY)
    Znp = np.array(Z)
    dZnp = np.array(dZ)
    var_dict["Z"] = Znp
    var_dict["X"] = dXnp
    var_dict["Y"] = dYnp
    var_dict["dZ"] = dZnp
    var_dict["Shim Pack/Change Step"] = adjustment_steps

    with outwidget:
        print("Survey Data File is Read for Subsequent Shimming Adjustments")

    return var_dict


# %%


def save_to_excel(modulename, axis, option, shim_numbers, datadict, shim_step):
    if int(shim_step[0]) == 0:
        save_to_excel_it1(
            modulename, axis, option, shim_numbers, datadict, shim_step[0]
        )
    if int(shim_step[0]) > 0:
        save_to_excel_it2(
            modulename, axis, option, shim_numbers, datadict, shim_step[0]
        )


def generate_output_filename(modulename, axis, num_iteration):
    """ Generates the output excel filename based on the module and the iteration """

    date = datetime.now().strftime("%Y%m%d%H%M%S")
    num_iteration_str = str(num_iteration).zfill(2)
    path_to_folder = str(Path(os.getcwd()).parent.absolute().parent.absolute())+'/'+modulename[0:4]+'/'+modulename
    if num_iteration == 1:
        filename = (
            path_to_folder
            + "/"
            + modulename
            + "_"
            + axis
            + "_module_shim_pack_"
            + num_iteration_str
            + "_"
            + date
            + ".xlsx"
        )
    else:
        filename = (
            path_to_folder
            + "/"
            + modulename
            + "_"
            + axis
            + "_module_shim_change_"
            + num_iteration_str
            + "_"
            + date
            + ".xlsx"
        )
    try:
        os.mkdir(path_to_folder)
    except FileExistsError:
        pass
#     excel_output_dir = Path(modulename)
#     excel_filename = str(excel_output_dir / filename)

    return filename


def save_to_excel_it1(modulename, axis, option, shim_numbers, datadict, num_iteration):

    indextitle = modulename + " Magnets"
    table = {
        indextitle: datadict["Magnet"],
        "QR Code Number": datadict["QR_Code_Num"],
        "Original dX": datadict["X"],
        "Original dY": datadict["Y"],
        "Z": datadict["Z"],
    }
    data = pd.DataFrame(table)
    data.set_index(indextitle)
    shim_option = shim_numbers[option - 1]
    
    data['Shim Change (0.001")'] = shim_option
    data['Total Shims (0.001")'] = shim_option
    #    data["Shim Axis = " + axis] = ""
    #    data["Standard error weight = " + str(stderr_weight.value)] = ""
    #    data["Shim calculation bounds = " + str(shim_range_slider.value)] = ""
    #    data["Selected Iteration "+str(num_iteration)+" Option = " + str(option)] = ""
    data["Shim Axis"] = axis
    data["Standard error weight"] = str(stderr_weight.value)
    data["Shim calculation bounds"] = str(shim_range_slider.value)
    data["Shim Pack/Change Step"] = str(num_iteration)
    data["Selected Option"] = str(option)
    filename = generate_output_filename(modulename, axis, num_iteration)
    #    filename = modulename+"_"+axis+"_Shimmed_It"+str(num_iteration)+".xlsx"
    data.to_excel(filename, index=False)
    
    if modulename[0:3] == 'QMQ':
        q_indices = set([i for i in range(len(datadict["Magnet"])) if 'Q' in datadict["Magnet"][i]])
        m_index = list((set(range(len(datadict["Magnet"]))) - q_indices))[0]
        if abs(datadict[axis][m_index]) > 200e-6:
            highlight_cells(filename)
    
    with widget_save_out_it1:
        print(
            "Option "
            + str(option)
            + " of initial shimming pack successfully saved to "
            + filename
        )
        display(data)

def highlight_cells(workbook):
    wb = load_workbook(workbook)
    ws = wb.active
    for col in [6,7]:
        for row in [3]:
            ws.cell(row, col).fill = PatternFill(start_color='00ffff0b', end_color='00ffff0b', fill_type = "solid")
    wb.save(workbook)

def save_to_excel_it2(modulename, axis, option, shim_numbers, datadict, num_iteration):

    indextitle = modulename + " Magnets"
    table = {
        indextitle: datadict["Magnet"],
        "QR Code Number": datadict["QR_Code_Num"],
        "Original dX": datadict["X"],
        "Original dY": datadict["Y"],
        "Z": datadict["Z"],
    }
    data = pd.DataFrame(table)
    data.set_index(indextitle)
    shim_option = shim_numbers[option - 1]
    data['Shim Change (0.001")'] = shim_option
    data['Total Shims (0.001")'] = shim_option + datadict["Current Shim Values"]
    #    data["Shim Axis = " + axis] = ""
    #    data["Standard error weight = " + str(stderr_weight.value)] = ""
    #    data["Shim calculation bounds = " + str(shim_range_slider.value)] = ""
    #    data["Selected Iteration "+str(num_iteration)+" Option = " + str(option)] = ""
    data["Shim Axis"] = axis
    data["Standard error weight"] = str(stderr_weight.value)
    data["Shim calculation bounds"] = str(shim_range_slider.value)
    data["Shim Pack/Change Step"] = str(num_iteration)
    data["Selected Option"] = str(option)
    filename = generate_output_filename(modulename, axis, num_iteration)
    #    filename = modulename+"_"+axis+"_Shimmed_It"+str(num_iteration)+".xlsx"
    data.to_excel(filename, index=False)
    with widget_save_out_it2:
        print(
            "Option "
            + str(option)
            + " of shimming change calculation successfully saved to "
            + filename
        )
        display(data)

def save_to_excel_dZ(modulename, axis, option, shim_numbers, datadict, num_iteration):

    indextitle = modulename + " Magnets"
    
    data = extract_shim_data_it2(
        module.value, fc_it2.selected, fc.selected, widget_save_out_dZ
    )
    dZ_origin = data["dZ"]
    z_move = []
    for i, magnet in enumerate(datadict["Magnet"]):
        if 'M' in magnet:
            z_move.append(-dZ_origin[i])
        else:
            z_move.append(0)
    
    table = {
        indextitle: datadict["Magnet"],
        "QR Code Number": datadict["QR_Code_Num"],
        "Original dX": datadict["X"],
        "Original dY": datadict["Y"],
        "Original dZ": dZ_origin,
        "Z": datadict["Z"],
        "Z Move [m]": z_move
    }
    data = pd.DataFrame(table)
    data.set_index(indextitle)
#     shim_option = shim_numbers[option - 1]
#     data['Shim Change (0.001")'] = shim_option
#     data['Total Shims (0.001")'] = shim_option + datadict["Current Shim Values"]
    #    data["Shim Axis = " + axis] = ""
    #    data["Standard error weight = " + str(stderr_weight.value)] = ""
    #    data["Shim calculation bounds = " + str(shim_range_slider.value)] = ""
    #    data["Selected Iteration "+str(num_iteration)+" Option = " + str(option)] = ""
    data["Shim Axis"] = axis
#     data["Standard error weight"] = str(stderr_weight.value)
#     data["Shim calculation bounds"] = str(shim_range_slider.value)
#     data["Shim Pack/Change Step"] = str(num_iteration)
#     data["Selected Option"] = str(option)
    filename = generate_output_filename(modulename, axis, num_iteration)
    #    filename = modulename+"_"+axis+"_Shimmed_It"+str(num_iteration)+".xlsx"
    data.to_excel(filename, index=False)
    with widget_save_out_dZ:
        print(
            "Option "
            + str(option)
            + " of shimming change calculation successfully saved to "
            + filename
        )
        display(data)

# %%
def shim_calculation_it1(
    name_module,
    axis,
    search_range=[-1, 3],
    stderr_weight=1.0,
    num_options=5,
    plot_bool=False,
):
    if name_module[0:3] == "DLM" or name_module[0:3] == "FOD":
        return shim_calculation_it1_DLMFODO(name_module, axis, search_range, stderr_weight, num_options, plot_bool)
    elif name_module[0:3] == "QMQ":
        return shim_calculation_it1_QMQ(name_module, axis, search_range, stderr_weight, num_options, plot_bool)
    
# %%
def shim_calculation_it1_QMQ(
    name_module,
    axis,
    search_range=[-1, 3],
    stderr_weight=1.0,
    num_options=5,
    plot_bool=False,
):
    """shim_calculation_it1 determines the shim magnitudes required to best align an arbitrary 
    number of magnet centers using the Pool multiprocessing module. Used for the first iteration 
    of shimming. Returns array of best shim numbers, array of shimmed displacements, and array 
    of standard errors. Writes best shim number arrays to a csv file named 
    name_module_axis_Shimmed_It1.csv.


    Args:
        name_module (str): name of the desired module file Z [mm], dX [mm], dY [mm]
        axis (str): the axis the shim calculation is performed on "X","Y","Z"
        search_range (list, optional): Search range for shim numbers below the offset Defaults to [-1, 3].
        stderr_weight (float, optional): Weight of Std Error from 0 to 1. Defaults to 1.0.
        num_options (int, optional): Number of displayed best shimming options. Defaults to 5.
        plot_bool (bool, optional): Plotting enabled. Defaults to False.

    Returns:
        list: List containing results of the calculation.
    """

    tic = time.perf_counter()
    with widget_stdout_it1:
        print(
            "\nStarting Initial Shim Pack Calcuation for "
            + name_module
            + " in the "
            + axis
            + " axis with a standard error weight of "
            + str(stderr_weight)
            + ".\n"
        )

    #  filename_module = "CDBInfo/" + module.value + ".xlsx" # extract_from_cdb(module.value) # !FUNCTION WIP!
    modulevalue = module.value
    filename_module = moduleToCDBInfoFilename(modulevalue)
    refreshMagnetConfigurationDataFromCDBIfPossible(filename_module, widget_stdout_it1)
    data = extract_shim_data_it1(
        name_module, filename_module, fc.selected, widget_stdout_it1
    )
    if data == None:
        with widget_stdout_it1:
            print("Error Reading Shim 1 Data, Calculation Stopped")
        return
    crd_name = data["Magnet"]
    qr_code = data["QR_Code_Num"]
    Z_np = data["Z"]
    X = data["X"]
    Y = data["Y"]
    if axis == "X":
        X_np = X
    else:
        X_np = Y
    num_pts = len(Z_np)
    MagShim_np = np.ones(num_pts) * 25e-6  # shim magnitudes: increments of 25 microns
   
    bestfit = []
    q_indices = [i for i in range(len(crd_name)) if 'Q' in crd_name[i]]
    max_dX = max([X_np[i] for i in q_indices]) 
    q_any = any([abs(X_np[i]) > 100e-6 for i in q_indices])
    for i in range(len(X_np)):
        if i in q_indices:
            if q_any: 
                if X_np[i] == max_dX:
                    bestfit.append(0)
                else:
                    q_indices.remove(i)
                    floor = np.floor((X_np[q_indices[0]] - X_np[i]) * (1 / 25e-6))
                    ceil = np.ceil((X_np[q_indices[0]] - X_np[i]) * (1 / 25e-6))
                    if (X_np[q_indices[0]] - (X_np[i] + 25e-6*floor)) > (X_np[i] + 25e-6*ceil - X_np[q_indices[0]]):
                        bestfit.append(int(ceil))
                    else:
                        bestfit.append(int(floor))
            else: 
                bestfit.append(0)
        else:
            bestfit.append(0)
    
    X_tofit = X_np + np.multiply(bestfit, MagShim_np)
    regression_result = scistats.linregress(Z_np, X_tofit)
    stderr = float(regression_result.stderr)

    best_stderr = [stderr]
    bestfit_vector = [bestfit]

    X_best = [
        X_np + np.multiply(i, MagShim_np) for i in bestfit_vector
    ]  # X_best is shimmed deviation (dX or dY) array
    
    q_indices = set([i for i in range(len(crd_name)) if 'Q' in crd_name[i]])
    m_index = list((set(range(len(crd_name))) - q_indices))[0]
    X_q = [X_tofit[i] for i in q_indices]
    Z_q = [Z_np[i] for i in q_indices]
    fx = np.polyfit(Z_q, X_q, 1)
    fx_func = np.poly1d(fx)
    
    m_optimal = fx_func(Z_np[m_index])
    m_offset = m_optimal-X_np[m_index]
    bestfit_vector[0][m_index] = "Move " + str(round(m_offset*1e6,3)) + " microns"

    toc = time.perf_counter()
    with widget_stdout_it1:
        print(
            "Elapsed time =",
            "{:.2f}".format(toc - tic),
            "s or",
            "{:.2f}".format((toc - tic) / 60),
            " mins",
        )

    for i in range(len(best_stderr)):
        with widget_stdout_it1:
            X_temp = X_best[i]
            X_temp[m_index] = X_np[m_index]
            z1 = np.polyfit(Z_np, X_temp, 1)
            p1 = np.poly1d(z1)
            diff = X_best[i] - p1(Z_np)
            rmse = math.sqrt(np.square(diff).mean())  # root mean square error
            print(
                "Shim numbers:",
                bestfit_vector[i],
                "| root mean square error:",
                "{:.2f}".format(rmse / 1e-6),
                "microns | improved dX:",
                X_best[i] / 1e-6,
                "microns | distance from trendline:",
                diff / 1e-6,
                "microns",
            )

    if plot_bool:
        for i in range(len(best_stderr)):
            with widget_stdout_it1:
                plot(Z_np, X_np, X_best[i], axis, name_module + " Option " + str(i + 1))

    with widget_stdout_it1:
        print("-------------------------------")
        print("Input Data For Calculation ")
        print("                 " + module.value)
        datatodisplay = copy.deepcopy(data)
        module_column = module.value + " Magnet"
        datatodisplay[module_column] = data["Magnet"]
        datatodisplay["QR Code"] = datatodisplay["QR_Code_Num"]
        datatodisplay["dX [m]"] = data["X"]
        datatodisplay["dY [m]"] = data["Y"]
        datatodisplay["Z [m]"] = data["Z"]
        datatodisplay.pop("Magnet")
        datatodisplay.pop("X")
        datatodisplay.pop("Y")
        datatodisplay.pop("Z")
        datatodisplay.pop("QR_Code_Num")
        displaypd = pd.DataFrame(
            datatodisplay,
            columns=[
                module_column,
                "QR Code",
                "dX [m]",
                "dY [m]",
                "Z [m]",
                "Shim Pack/Change Step",
            ],
        )
        display(displaypd)

    return (
        modulevalue,
        axis,
        data,
        bestfit_vector,
        X_best,
        best_stderr,
        data["Shim Pack/Change Step"],
    )

# %%
def shim_calculation_it1_DLMFODO(
    name_module,
    axis,
    search_range=[-1, 3],
    stderr_weight=1.0,
    num_options=5,
    plot_bool=False,
):
    """shim_calculation_it1 determines the shim magnitudes required to best align an arbitrary 
    number of magnet centers using the Pool multiprocessing module. Used for the first iteration 
    of shimming. Returns array of best shim numbers, array of shimmed displacements, and array 
    of standard errors. Writes best shim number arrays to a csv file named 
    name_module_axis_Shimmed_It1.csv.


    Args:
        name_module (str): name of the desired module file Z [mm], dX [mm], dY [mm]
        axis (str): the axis the shim calculation is performed on "X","Y","Z"
        search_range (list, optional): Search range for shim numbers below the offset Defaults to [-1, 3].
        stderr_weight (float, optional): Weight of Std Error from 0 to 1. Defaults to 1.0.
        num_options (int, optional): Number of displayed best shimming options. Defaults to 5.
        plot_bool (bool, optional): Plotting enabled. Defaults to False.

    Returns:
        list: List containing results of the calculation.
    """

    tic = time.perf_counter()
    with widget_stdout_it1:
        print(
            "\nStarting Initial Shim Pack Calcuation for "
            + name_module
            + " in the "
            + axis
            + " axis with a standard error weight of "
            + str(stderr_weight)
            + ".\n"
        )

    #  filename_module = "CDBInfo/" + module.value + ".xlsx" # extract_from_cdb(module.value) # !FUNCTION WIP!
    modulevalue = module.value
    filename_module = moduleToCDBInfoFilename(modulevalue)
    refreshMagnetConfigurationDataFromCDBIfPossible(filename_module, widget_stdout_it1)
    data = extract_shim_data_it1(
        name_module, filename_module, fc.selected, widget_stdout_it1
    )
    if data == None:
        with widget_stdout_it1:
            print("Error Reading Shim 1 Data, Calculation Stopped")
        return
    crd_name = data["Magnet"]
    qr_code = data["QR_Code_Num"]
    Z_np = data["Z"]
    X = data["X"]
    Y = data["Y"]
    if axis == "X":
        X_np = X
    else:
        X_np = Y
    num_pts = len(Z_np)
    MagShim_np = np.ones(num_pts) * 25e-6  # shim magnitudes: increments of 25 microns

    max_dX = max(X_np)
    deviation_offset = np.array(max_dX - X_np)
    ShimMid = np.floor(deviation_offset * (1 / 25e-6)).astype(int)
    ShimStart = ShimMid + search_range[0]
    ShimEnd = ShimMid + search_range[1]
    ShimStart = np.maximum(0, ShimStart)

    ShimRange = []
    for index_pt in range(num_pts):
        ShimRange.append(range(ShimStart[index_pt], ShimEnd[index_pt]))

    permutations = list(itertools.product(*ShimRange))
    permutations = [i for i in permutations if zero_in_set(i)]
    num_permutations = len(permutations)
    with widget_stdout_it1:
        print("Number of permutations:", num_permutations)
    index_permutations = range(0, num_permutations)

    num_cores = os.cpu_count()
    p = Pool(
        processes=num_cores,
        initializer=init_worker,
        initargs=(permutations, X_np, Z_np, MagShim_np, stderr_weight),
    )
    X_bestfit_data = p.map(lin_regress, index_permutations)
    p.close()
    p.join()

    num_best = num_options
    stderr, weighted_score = (
        np.array(X_bestfit_data)[:, 0],
        np.array(X_bestfit_data)[:, 1],
    )
    argmin_bestfit_values = np.argsort(weighted_score)[:num_best]
    best_stderr, bestfit_vector = (
        stderr[argmin_bestfit_values],
        np.array(permutations)[argmin_bestfit_values, :],
    )  # best_stderr is standard error array and bestfit_vector is shim number array
    X_best = [
        X_np + np.multiply(i, MagShim_np) for i in bestfit_vector
    ]  # X_best is shimmed deviation (dX or dY) array

    toc = time.perf_counter()
    with widget_stdout_it1:
        print(
            "Elapsed time =",
            "{:.2f}".format(toc - tic),
            "s or",
            "{:.2f}".format((toc - tic) / 60),
            " mins",
        )

    for i in range(len(best_stderr)):
        with widget_stdout_it1:
            z1 = np.polyfit(Z_np, X_best[i], 1)
            p1 = np.poly1d(z1)
            diff = X_best[i] - p1(Z_np)
            rmse = math.sqrt(np.square(diff).mean())  # root mean square error
            print(
                "Shim numbers:",
                bestfit_vector[i],
                "| root mean square error:",
                "{:.2f}".format(rmse / 1e-6),
                "microns | improved dX:",
                X_best[i] / 1e-6,
                "microns | distance from trendline:",
                diff / 1e-6,
                "microns",
            )

    if plot_bool:
        for i in range(len(best_stderr)):
            with widget_stdout_it1:
                plot(Z_np, X_np, X_best[i], axis, name_module + " Option " + str(i + 1))

    with widget_stdout_it1:
        print("-------------------------------")
        print("Input Data For Calculation ")
        print("                 " + module.value)
        datatodisplay = copy.deepcopy(data)
        module_column = module.value + " Magnet"
        datatodisplay[module_column] = data["Magnet"]
        datatodisplay["QR Code"] = datatodisplay["QR_Code_Num"]
        datatodisplay["dX [m]"] = data["X"]
        datatodisplay["dY [m]"] = data["Y"]
        datatodisplay["Z [m]"] = data["Z"]
        datatodisplay.pop("Magnet")
        datatodisplay.pop("X")
        datatodisplay.pop("Y")
        datatodisplay.pop("Z")
        datatodisplay.pop("QR_Code_Num")
        displaypd = pd.DataFrame(
            datatodisplay,
            columns=[
                module_column,
                "QR Code",
                "dX [m]",
                "dY [m]",
                "Z [m]",
                "Shim Pack/Change Step",
            ],
        )
        display(displaypd)

    return (
        modulevalue,
        axis,
        data,
        bestfit_vector,
        X_best,
        best_stderr,
        data["Shim Pack/Change Step"],
    )


#
def read_shim_data(shimDataFile):
    """ Attempts to read the shim data from the excel file and then returns a 
        numpy array with the existing shim numbers """
    wb_data_it1 = load_workbook(shimDataFile, data_only=True)
    ws = wb_data_it1.active
    col_names = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    NegShim = []
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):  
        if type(row_cells[col_names.index('Total Shims (0.001")')].value) is int or type(row_cells[col_names.index('Total Shims (0.001")')].value) is float:
            NegShim.append(-row_cells[col_names.index('Total Shims (0.001")')].value)
        else:
            NegShim.append(0)
    NegShim = np.array(NegShim)
    return NegShim

# %%
def shim_calculation_it2(
    name_module,
    axis,
    shimDataFile,
    search_range=[-1, 3],
    stderr_weight=1.0,
    num_options=5,
    plot_bool=False,
):
    if name_module[0:3] == "DLM" or name_module[0:3] == "FOD":
        return shim_calculation_it2_DLMFODO(name_module, axis, shimDataFile, search_range, stderr_weight, num_options, plot_bool)
    elif name_module[0:3] == "QMQ":
        return shim_calculation_it2_QMQ(name_module, axis, shimDataFile, search_range, stderr_weight, num_options, plot_bool)

# %%
def shim_calculation_it2_DLMFODO(
    name_module,
    axis,
    shimDataFile,
    search_range=[-1, 3],
    stderr_weight=1.0,
    num_options=5,
    plot_bool=False,
):

    """
    shim_calculation_it2 determines the shim magnitudes required to best align an arbitrary number of magnet centers using the Pool multiprocessing module. Used for the second iteration of shimming. Returns array of best shim numbers, array of shimmed displacements, and array of standard errors. Writes best shim number arrays to a csv file named name_module_axis_Shimmed_It2.csv.

    :param str name_module: name of the desired module to shim
    :param str axis: name of the axis the shim calculation is performed on, write "X" 
        for the x-axis and "Y" for the y-axis   
    :param int list search_range: search_range[0] is the desired shim numbers to search 
        below the offset and search_range[1] is the desired shim numbers to search above 
        the offset
    :param float stderr_weight: weight of the standard error, 1.0 is the highest,
        0.0 is the lowest, default is 1.0, weight of the total shim magnitude is 
        (1 - stderr_weight), although not required, it is suggested that stderr_weight 
        is within 0.7 to 1.0
    :param int num_options: number of best shimming options to be displayed and
        written to the new file
    :param bool plot_bool: True if plotting is enabled and False if plotting is disabled 

    """

    tic = time.perf_counter()
    with widget_stdout_it2:
        print(
            "\nStarting Shim Change Calculation for "
            + name_module
            + " in the "
            + axis
            + " axis with a standard error weight of "
            + str(stderr_weight)
            + ".\n"
        )

    modulevalue = module.value
    filename_module = moduleToCDBInfoFilename(modulevalue)
    data = extract_shim_data_it2(
        modulevalue, fc_it2.selected, fc.selected, widget_stdout_it2
    )
    if data == None:
        with widget_stdout_it2:
            print("Error Reading Survey for Subsequent Shimming Adjustments")
        return
    crd_name = data["Magnet"]
    qr_code = data["QR_Code_Num"]
    Z_np = data["Z"]
    X = data["X"]
    Y = data["Y"]
    if axis == "X":
        X_np = X
    else:
        X_np = Y
    num_pts = len(Z_np)
    MagShim_np = np.ones(num_pts) * 25e-6  # shim magnitudes: increments of 25 microns

    #    wb_data_it1 = load_workbook(name_module + "_" + axis + "_Shimmed_It1.xlsx", data_only=True)
    #    try:
    #        wb_data_it1 = load_workbook(shimDataFile, data_only=True)
    #    except:
    #        with widget_stdout_it2:
    #            print("Error Reading Shim Data for Iteration 2")
    #        return
    #    ws = wb_data_it1.active
    #    col_names = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    #    NegShim = np.array([-row_cells[col_names.index('Total Shims (0.001")')].value for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row)])
    try:
        NegShim = read_shim_data(shimDataFile)
    except:
        with widget_stdout_it2:
            print("Error Reading Shim Data for Subsequent Shimming Calculations")
        return
    with widget_stdout_it2:
        print("All Data Read")
    #    Ensure that the number of shims that we have match the number of magnets read in from
    #    the data file array
    if len(NegShim) != len(Z_np):
        with widget_stdout_it2:
            print("ERROR: NUMBER OF MAGNETS DOESN'T MATCH IN INPUT FILES ")
            print(
                "Number of Shims: ",
                len(NegShim),
                " Length of Centers Info: ",
                len(Z_np),
            )
        return

    with widget_stdout_it2:
        print("Lengths OK")
    max_dX = max(X_np)
    deviation_offset = np.array(max_dX - X_np)
    ShimMid = np.floor(deviation_offset * (1 / 25e-6)).astype(int)
    axis_shift = min(ShimMid - NegShim)
    ShimMid = ShimMid - axis_shift
    ShimStart = ShimMid + search_range[0]
    ShimEnd = ShimMid + search_range[1]
    ShimStart = np.maximum(NegShim, ShimStart)
    with widget_out:
        print("Shim Calc OK")

    ShimRange = []
    for index_pt in range(num_pts):
        ShimRange.append(range(ShimStart[index_pt], ShimEnd[index_pt]))

    permutations = list(itertools.product(*ShimRange))
    permutations = [i for i in permutations if neg_in_set(i, NegShim)]
    num_permutations = len(permutations)
    with widget_stdout_it2:
        print("Number of permutations:", num_permutations)
    index_permutations = range(0, num_permutations)

    num_cores = os.cpu_count()
    p = Pool(
        processes=num_cores,
        initializer=init_worker,
        initargs=(permutations, X_np, Z_np, MagShim_np, stderr_weight),
    )
    X_bestfit_data = p.map(lin_regress, index_permutations)
    p.close()
    p.join()

    num_best = num_options
    stderr, weighted_score = (
        np.array(X_bestfit_data)[:, 0],
        np.array(X_bestfit_data)[:, 1],
    )
    argmin_bestfit_values = np.argsort(weighted_score)[:num_best]
    best_stderr, bestfit_vector = (
        stderr[argmin_bestfit_values],
        np.array(permutations)[argmin_bestfit_values, :],
    )  # best_stderr is standard error array and bestfit_vector is shim number array
    X_best = [
        X_np + np.multiply(i, MagShim_np) for i in bestfit_vector
    ]  # X_best is shimmed deviation (dX or dY) array

    toc = time.perf_counter()
    with widget_stdout_it2:
        print(
            "Elapsed time =",
            "{:.2f}".format(toc - tic),
            "s or",
            "{:.2f}".format((toc - tic) / 60),
            " mins",
        )

    for i in range(len(best_stderr)):
        with widget_stdout_it2:
            z1 = np.polyfit(Z_np, X_best[i], 1)
            p1 = np.poly1d(z1)
            diff = X_best[i] - p1(Z_np)
            rmse = math.sqrt(np.square(diff).mean())  # root mean square error
            print(
                "Shim numbers:",
                bestfit_vector[i],
                "| root mean square error:",
                "{:.2f}".format(rmse / 1e-6),
                "microns | improved dX:",
                X_best[i] / 1e-6,
                "microns | distance from trendline:",
                diff / 1e-6,
                "microns",
            )

    if plot_bool:
        for i in range(len(best_stderr)):
            with widget_stdout_it2:
                plot(Z_np, X_np, X_best[i], axis, name_module + " Option " + str(i + 1))

    with widget_stdout_it2:
        print("-------------------------------")
        print("Input Data For Calculation ")
        print("                 " + module.value)
        datatodisplay = copy.deepcopy(data)
        module_column = module.value + " Magnet"
        datatodisplay[module_column] = data["Magnet"]
        datatodisplay["QR Code"] = datatodisplay["QR_Code_Num"]
        datatodisplay["dX"] = data["X"]
        datatodisplay["dY"] = data["Y"]
        datatodisplay["Current Shim Values"] = -NegShim
        datatodisplay.pop("Magnet")
        datatodisplay.pop("X")
        datatodisplay.pop("Y")
        displaypd = pd.DataFrame(
            datatodisplay,
            columns=[
                module_column,
                "QR Code",
                "dX [m]",
                "dY [m]",
                "Z [m]",
                "Shim Pack/Change Step",
            ],
        )
        displaypd = pd.DataFrame(
            datatodisplay,
            columns=[module_column, "QR Code", "dX", "dY", "Z", "Current Shim Values"],
        )
        display(displaypd)

    # Add the original shimming to the data array
    data["Current Shim Values"] = -NegShim
    return (
        modulevalue,
        axis,
        data,
        bestfit_vector,
        X_best,
        best_stderr,
        data["Shim Pack/Change Step"],
    )

# %%
def shim_calculation_it2_QMQ(
    name_module,
    axis,
    shimDataFile,
    search_range=[-1, 3],
    stderr_weight=1.0,
    num_options=5,
    plot_bool=False,
):

    """
    shim_calculation_it2 determines the shim magnitudes required to best align an arbitrary number of magnet centers using the Pool multiprocessing module. Used for the second iteration of shimming. Returns array of best shim numbers, array of shimmed displacements, and array of standard errors. Writes best shim number arrays to a csv file named name_module_axis_Shimmed_It2.csv.

    :param str name_module: name of the desired module to shim
    :param str axis: name of the axis the shim calculation is performed on, write "X" 
        for the x-axis and "Y" for the y-axis   
    :param int list search_range: search_range[0] is the desired shim numbers to search 
        below the offset and search_range[1] is the desired shim numbers to search above 
        the offset
    :param float stderr_weight: weight of the standard error, 1.0 is the highest,
        0.0 is the lowest, default is 1.0, weight of the total shim magnitude is 
        (1 - stderr_weight), although not required, it is suggested that stderr_weight 
        is within 0.7 to 1.0
    :param int num_options: number of best shimming options to be displayed and
        written to the new file
    :param bool plot_bool: True if plotting is enabled and False if plotting is disabled 

    """

    tic = time.perf_counter()
    with widget_stdout_it2:
        print(
            "\nStarting Shim Change Calculation for "
            + name_module
            + " in the "
            + axis
            + " axis with a standard error weight of "
            + str(stderr_weight)
            + ".\n"
        )

    modulevalue = module.value
    filename_module = moduleToCDBInfoFilename(modulevalue)
    data = extract_shim_data_it2(
        modulevalue, fc_it2.selected, fc.selected, widget_stdout_it2
    )
    if data == None:
        with widget_stdout_it2:
            print("Error Reading Survey for Subsequent Shimming Adjustments")
        return
    crd_name = data["Magnet"]
    qr_code = data["QR_Code_Num"]
    Z_np = data["Z"]
    X = data["X"]
    Y = data["Y"]
    if axis == "X":
        X_np = X
    else:
        X_np = Y
    num_pts = len(Z_np)
    MagShim_np = np.ones(num_pts) * 25e-6  # shim magnitudes: increments of 25 microns

    #    wb_data_it1 = load_workbook(name_module + "_" + axis + "_Shimmed_It1.xlsx", data_only=True)
    #    try:
    #        wb_data_it1 = load_workbook(shimDataFile, data_only=True)
    #    except:
    #        with widget_stdout_it2:
    #            print("Error Reading Shim Data for Iteration 2")
    #        return
    #    ws = wb_data_it1.active
    #    col_names = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    #    NegShim = np.array([-row_cells[col_names.index('Total Shims (0.001")')].value for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row)])
    try:
        NegShim = read_shim_data(shimDataFile)
    except:
        with widget_stdout_it2:
            print("Error Reading Shim Data for Subsequent Shimming Calculations")
        return
    with widget_stdout_it2:
        print("All Data Read")
    #    Ensure that the number of shims that we have match the number of magnets read in from
    #    the data file array
    if len(NegShim) != len(Z_np):
        with widget_stdout_it2:
            print("ERROR: NUMBER OF MAGNETS DOESN'T MATCH IN INPUT FILES ")
            print(
                "Number of Shims: ",
                len(NegShim),
                " Length of Centers Info: ",
                len(Z_np),
            )
        return

    with widget_stdout_it2:
        print("Lengths OK")
    
#     ShimRange = []
#     max_dX = max(X_np)
#     for i in range(len(X_np)):
#         if X_np[i] == max_dX:
#             ShimRange.append([0])
#         else:
#             floor = np.floor((max_dX - X_np[i]) * (1 / 25e-6))
#             ceil = np.ceil((max_dX - X_np[i]) * (1 / 25e-6))
#             if floor == ceil: 
#                 ShimRange.append([int(floor)])
#             else:
#                 ShimRange.append([int(floor),int(ceil)])
    
    q_indices = set([i for i in range(len(crd_name)) if 'Q' in crd_name[i]])
    m_index = list((set(range(len(crd_name))) - q_indices))[0]
    X_q = [X_np[i] for i in q_indices]
    Z_q = [Z_np[i] for i in q_indices]
    fx = np.polyfit(Z_q, X_q, 1)
    fx_func = np.poly1d(fx)
    
    m_optimal = fx_func(Z_np[m_index])
    ceil = np.ceil((m_optimal-X_np[m_index])/25e-6)
    floor = np.floor((m_optimal-X_np[m_index])/25e-6)

    if (ceil*25e-6 + X_np[m_index] - m_optimal) > -(floor*25e-6 + X_np[m_index] - m_optimal):
        m_shim = floor
    else:
        m_shim = ceil
       
    bestfit = []
    if X_np[m_index] < X_np[list(q_indices)[0]]:
        for i in range(len(X_np)):
            if i == m_index:
                bestfit.append(m_shim)
            else:
                bestfit.append(0)
    else:
        for i in range(len(X_np)):
            if i == m_index:
                bestfit.append(0)
            else:
                bestfit.append(-m_shim)
            
    X_tofit = X_np + np.multiply(bestfit, MagShim_np)
    regression_result = scistats.linregress(Z_np, X_tofit)
    stderr = float(regression_result.stderr)

    best_stderr = [stderr]
    bestfit_vector = [bestfit]
    
    X_best = [
        X_np + np.multiply(i, MagShim_np) for i in bestfit_vector
    ]  # X_best is shimmed deviation (dX or dY) array

    toc = time.perf_counter()
    with widget_stdout_it2:
        print(
            "Elapsed time =",
            "{:.2f}".format(toc - tic),
            "s or",
            "{:.2f}".format((toc - tic) / 60),
            " mins",
        )

    for i in range(len(best_stderr)):
        with widget_stdout_it2:
            z1 = np.polyfit(Z_np, X_best[i], 1)
            p1 = np.poly1d(z1)
            diff = X_best[i] - p1(Z_np)
            rmse = math.sqrt(np.square(diff).mean())  # root mean square error
            print(
                "Shim numbers:",
                bestfit_vector[i],
                "| root mean square error:",
                "{:.2f}".format(rmse / 1e-6),
                "microns | improved dX:",
                X_best[i] / 1e-6,
                "microns | distance from trendline:",
                diff / 1e-6,
                "microns",
            )

    if plot_bool:
        for i in range(len(best_stderr)):
            with widget_stdout_it2:
                plot(Z_np, X_np, X_best[i], axis, name_module + " Option " + str(i + 1))

    with widget_stdout_it2:
        print("-------------------------------")
        print("Input Data For Calculation ")
        print("                 " + module.value)
        datatodisplay = copy.deepcopy(data)
        module_column = module.value + " Magnet"
        datatodisplay[module_column] = data["Magnet"]
        datatodisplay["QR Code"] = datatodisplay["QR_Code_Num"]
        datatodisplay["dX"] = data["X"]
        datatodisplay["dY"] = data["Y"]
        datatodisplay["Current Shim Values"] = -NegShim
        datatodisplay.pop("Magnet")
        datatodisplay.pop("X")
        datatodisplay.pop("Y")
        displaypd = pd.DataFrame(
            datatodisplay,
            columns=[
                module_column,
                "QR Code",
                "dX [m]",
                "dY [m]",
                "Z [m]",
                "Shim Pack/Change Step",
            ],
        )
        displaypd = pd.DataFrame(
            datatodisplay,
            columns=[module_column, "QR Code", "dX", "dY", "Z", "Current Shim Values"],
        )
        display(displaypd)

    # Add the original shimming to the data array
    data["Current Shim Values"] = -NegShim
    return (
        modulevalue,
        axis,
        data,
        bestfit_vector,
        X_best,
        best_stderr,
        data["Shim Pack/Change Step"],
    )

# %%
def on_button_clicked(b):

    """
    on_button_clicked is the callback function ran when button is clicked. Displays table containing coordinate names, QR code numbers, Z, dX, and dY.  

    """

    with output:
        with widget_out:
            clear_output(wait=False)
        if len(module.value) == 0:
            with widget_out:
                print(Fore.RED + "Please enter the module name." + Style.RESET_ALL)
        elif fc.selected == None:
            with widget_out:
                print(
                    Fore.RED
                    + "Please enter the filename of fiducials spreadsheet."
                    + Style.RESET_ALL
                )
        else:
            filename_module = moduleToCDBInfoFilename(module.value)
            with widget_out:
                print("Shim Calculation Version: ",VERSION)
                print("Number of CPU Cores on this machine is: ", os.cpu_count())
            refreshMagnetConfigurationDataFromCDBIfPossible(filename_module, widget_out)
            if fiducial_file_format.value == "First Iteration":
                data = extract_shim_data_it1(
                    module.value, filename_module, fc.selected, widget_out
                )
                with widget_out:
                    print("Extracting magnet assignment data from ", filename_module)
                    print("Magnet Offset/Fiducial data is in", fc.selected)

                if data == None:
                    with widget_out:
                        print(
                            "Error Reading Data Shim Data for Initial Shim Pack Calculation"
                        )
                    return
                with widget_out:
                    print(
                        "Data Extracted from ",
                        filename_module,
                        " and ",
                        fc.selected,
                        " using Initial Shim Pack Format",
                    )
            if fiducial_file_format.value == "Second Iteration":
                with widget_out:
                    print("Extracting magnet assignment data from ", fc_it2.selected)
                    print("Magnet Offset/Fidƒucial data is in", fc.selected)
                    data = extract_shim_data_it2(
                        module.value, fc_it2.selected, fc.selected, widget_out
                    )
                    print("Shim Data Extracted for Initial Shim Calculation")

                if data == None:
                    with widget_out:
                        print(
                            "Error Reading Data Shim Data for Subsquent Shim Calculation"
                        )
                    return
                with widget_out:
                    print(
                        "Data Extracted from ",
                        fc_it2.selected,
                        " and ",
                        fc.selected,
                        " using Shim Change Format",
                    )
                try:
                    NegShim = read_shim_data(fc_it2.selected)
                    data["Current Shim Values"] = -NegShim
                except:
                    with widget_out:
                        print("Error reading shim values from: ", fc_it2.selected)
                        return
            indextitle = module.value + " Magnets"
            with widget_out:
                # Lets make sure the length of the arrays are OK:
                length_of_data_arrays = {len(data[key]) for key in data.keys()}
                if len(length_of_data_arrays) != 1:
                    print("ERROR: NUMBER OF MAGNETS DOESN'T MATCH IN INPUT FILES ")
                    #FIXME: Added print for data to debug problem with FODO
                    print(data)
                print("Preparing Display Data Frames")
                try:
                    if fiducial_file_format.value == "First Iteration":
                        table = {
                            indextitle: data["Magnet"],
                            "QR Code Number": data["QR_Code_Num"],
                            "Z": data["Z"],
                            "dX": data["X"],
                            "dY": data["Y"],
                        }
                    elif fiducial_file_format.value == "Second Iteration":
                        table = {
                            indextitle: data["Magnet"],
                            "QR Code Number": data["QR_Code_Num"],
                            "Z": data["Z"],
                            "dX": data["X"],
                            "dY": data["Y"],
                            'Current Shim Values (0.001")': data["Current Shim Values"],
                        }
                except Exception as e:
                    print(str(e))
                pd_to_display = pd.DataFrame(table)
                pd_to_display.set_index(indextitle)
                print("                 " + module.value)
                display(pd_to_display)


# %%
def on_button_clicked_it1(b):

    """
    on_button_clicked_it1 is the callback function ran when button_it1 is clicked. Executes shim_calculation_it1. 

    """

    with output_it1:
        with widget_stdout_it1:
            clear_output(wait=False)
        if len(module.value) == 0:
            with widget_stdout_it1:
                print(
                    Fore.RED + "Error: Please enter the module name." + Style.RESET_ALL
                )
        elif fc.selected == None:
            with widget_stdout_it1:
                print(
                    Fore.RED
                    + "Error: Please enter the filename of fiducials spreadsheet."
                    + Style.RESET_ALL
                )
        else:
            global shim_numbers_it1
            global module_it1
            global datadict_it1
            global axis_it1
            global shimming_step
            try:
                (
                    module_it1,
                    axis_it1,
                    datadict_it1,
                    shim_numbers_it1,
                    X_best_it1,
                    best_stderr_it1,
                    shimming_step,
                ) = shim_calculation_it1(
                    module.value,
                    axis.value,
                    search_range=shim_range_slider.value,
                    stderr_weight=stderr_weight.value,
                    num_options=num_options.value,
                    plot_bool=plot_bool.value,
                )
            except:
                with widget_stdout_it1:
                    print(Fore.RED + "Calculation Error and Aborted")


# %%
def on_button_clicked_savefile_it1(b):

    """
    on_button_clicked_savefile_it1 is the callback function ran when button_savefile_it1 is clicked. Write first iteration shimming data to an Excel file. 

    """

    with output_savefile_it1:
        with widget_save_out_it1:
            clear_output(wait=False)
        save_to_excel(
            module_it1,
            axis_it1,
            it1_option.value,
            shim_numbers_it1,
            datadict_it1,
            shimming_step,
        )


# %%
def on_button_clicked_it2(b):

    """
    on_button_clicked_it2 is the callback function ran when button_it2 is clicked. Executes shim_calculation_it2. 

    """

    with output_it2:
        with widget_stdout_it2:
            clear_output(wait=False)
        if len(module.value) == 0:
            with widget_stdout_it2:
                print(
                    Fore.RED + "Error: Please enter the module name." + Style.RESET_ALL
                )
        elif fc.selected == None:
            with widget_stdout_it2:
                print(
                    Fore.RED
                    + "Error: Please enter the filename of fiducials spreadsheet."
                    + Style.RESET_ALL
                )
        else:
            global shim_numbers_it2
            global module_it2
            global datadict_it2
            global axis_it2
            global shimming_step
            #            try:
            (
                module_it2,
                axis_it2,
                datadict_it2,
                shim_numbers_it2,
                X_best_it2,
                best_stderr_it2,
                shimming_step,
            ) = shim_calculation_it2(
                module.value,
                axis.value,
                fc_it2.selected,
                search_range=shim_range_slider.value,
                stderr_weight=stderr_weight.value,
                num_options=num_options.value,
                plot_bool=plot_bool.value,
            )


#            except:
#                with widget_stdout_it2:
#                    print(Fore.RED + "Calculation Error and Aborted")

# %%
def on_button_clicked_savefile_it2(b):

    """
    on_button_clicked_savefile_it2 is the callback function ran when button_savefile_it2 is clicked. Write second iteration shimming data to an Excel file. 

    """

    with output_savefile_it2:
        with widget_save_out_it2:
            clear_output(wait=False)
        save_to_excel(
            module_it2,
            axis_it2,
            it2_option.value,
            shim_numbers_it2,
            datadict_it2,
            shimming_step,
        )

def on_value_change(change):
    if fiducial_file_format.value == "Second Iteration":
        try:
            path_to_module = str(Path(os.getcwd()).parent.absolute().parent.absolute())+"/"+change['new'][0:4]+"/"+change['new']
            fc.reset(path=path_to_module)
            fc_it2.reset(path=path_to_module)
        except:
            pass
# %%

module = widgets.Dropdown(
    options=MagnetModuleNames,
    description="Module Name",
    disabled=False,
    style={"description_width": "initial"},
    layout=widgets.Layout(width="auto", height="auto"),
)
module.observe(on_value_change, names='value')
fiducial_file_format = widgets.ToggleButtons(
    options=["First Iteration", "Second Iteration"],
    description="Fiducial Spreadsheet Format: ",
    disabled=False,
    style={"description_width": "initial"},
    layout=widgets.Layout(width="auto", height="auto"),
)
# Get the Fiducials Filename
# Add a file chooser to get the fiducials filename
path_to_fiducials = str(Path(os.getcwd()).parent.absolute().parent.absolute())+"/MM_FIDUCIALS_ALL"
fc = FileChooser(path_to_fiducials)
fc.use_dir_icons = True
fc.filter_pattern = ["*.xlsx", "*.txt"]
fc.title = "Filename of Fiducials Spreadsheet or Final Survey CSV:"
fc.default_filename = "APS_U_MagnetFiducials_All.xlsx"


widget_out = widgets.Output(layout={"border": "1px solid black"})
button = widgets.Button(
    description="Extract and display input data for calculation",
    layout=widgets.Layout(width="auto", height="auto"),
)
buttonresetdefault = widgets.Button(
    description="Reset fiducials filename to default",
    layout=widgets.Layout(width="auto", height="auto"),
)
output = widgets.Output(height="100%")
button.on_click(on_button_clicked)
box = widgets.VBox([module, fiducial_file_format, fc, button, widget_out, output])

# %%

#
# Below are all of the widgets defined for the program.   Each notebook will simply render the ones that it needs
#

max_option = 20
np.set_printoptions(precision=1)

widget_stdout_it1 = widgets.Output(layout={"border": "1px solid black"})
widget_stdout_it2 = widgets.Output(layout={"border": "1px solid black"})
axis = widgets.ToggleButtons(options=["X", "Y"], description="Axis:", disabled=False)
it1_option = widgets.IntSlider(
    value=1,
    min=1,
    max=max_option,
    step=1,
    description="Selected shim pack option:",
    disabled=False,
    continuous_update=False,
    orientation="horizontal",
    readout=True,
    readout_format="d",
    style={"description_width": "initial"},
    layout=widgets.Layout(width="auto", height="auto"),
)
it2_option = widgets.IntSlider(
    value=1,
    min=1,
    max=max_option,
    step=1,
    description="Selected shim change option:",
    disabled=False,
    continuous_update=False,
    orientation="horizontal",
    readout=True,
    readout_format="d",
    style={"description_width": "initial"},
    layout=widgets.Layout(width="auto", height="auto"),
)
stderr_weight = widgets.FloatSlider(
    value=1.0,
    min=0,
    max=1.0,
    step=0.01,
    description="Standard error weight:",
    disabled=False,
    continuous_update=False,
    orientation="horizontal",
    readout=True,
    readout_format=".2f",
    style={"description_width": "initial"},
    layout=widgets.Layout(width="auto", height="auto"),
)
num_options = widgets.IntSlider(
    value=5,
    min=1,
    max=max_option,
    step=1,
    description="Number of options displayed:",
    disabled=False,
    continuous_update=False,
    orientation="horizontal",
    readout=True,
    readout_format="d",
    style={"description_width": "initial"},
    layout=widgets.Layout(width="auto", height="auto"),
)
plot_bool = widgets.Checkbox(
    value=True, description="Display plots", disabled=False, indent=False
)
shim_range_slider = widgets.IntRangeSlider(
    value=[-1, 3],
    min=-5,
    max=7,
    step=1,
    description="Shim calculation bounds:",
    disabled=False,
    continuous_update=False,
    orientation="horizontal",
    readout=True,
    readout_format="d",
    layout=widgets.Layout(width="auto", height="auto"),
    style={"description_width": "initial"},
)

button_it1 = widgets.Button(
    description="Run Shim Pack Calculation.",
    layout=widgets.Layout(width="auto", height="auto"),
)
output_it1 = widgets.Output()
button_it1.on_click(on_button_clicked_it1)

shim_numbers_it1 = np.array([])
widget_save_out_it1 = widgets.Output(layout={"border": "1px solid black"})
button_savefile_it1 = widgets.Button(
    description="Write shim pack result to an Excel file.",
    layout=widgets.Layout(width="auto", height="auto"),
)
output_savefile_it1 = widgets.Output()
button_savefile_it1.on_click(on_button_clicked_savefile_it1)
############### Second Iteration
button_it2 = widgets.Button(
    description="Run Shim Change Calculation.",
    layout=widgets.Layout(width="auto", height="auto"),
)
output_it2 = widgets.Output()
button_it2.on_click(on_button_clicked_it2)

path_to_main = str(Path(os.getcwd()).parent.absolute().parent.absolute())
fc_it2 = FileChooser(path_to_main)
fc_it2.use_dir_icons = True
fc_it2.filter_pattern = ["*.xlsx", "*.txt"]
fc_it2.title = "Filename containing Shim Values for Shimming Change:"

shim_numbers_it2 = np.array([])
widget_save_out_it2 = widgets.Output(layout={"border": "1px solid black"})
button_savefile_it2 = widgets.Button(
    description="Write shim change results to an Excel file.",
    layout=widgets.Layout(width="auto", height="auto"),
)
output_savefile_it2 = widgets.Output()
button_savefile_it2.on_click(on_button_clicked_savefile_it2)

box_savefile_it1 = widgets.VBox(
    [it1_option, button_savefile_it1, output_savefile_it1, widget_save_out_it1]
)
box_it1 = widgets.VBox(
    [
        axis,
        shim_range_slider,
        stderr_weight,
        num_options,
        plot_bool,
        button_it1,
        widget_stdout_it1,
        output_it1,
    ]
)
box_it2 = widgets.VBox([fc_it2, button_it2, widget_stdout_it2, output_it2])
box_savefile_it2 = widgets.VBox(
    [it2_option, button_savefile_it2, output_savefile_it2, widget_save_out_it2]
)

def on_button_clicked_savefile_dZ(b):
    
    with output_savefile_dZ:
        with widget_save_out_dZ:
            clear_output(wait=False)
        save_to_excel_dZ(
            module_it2,
            "Z",
            it2_option.value,
            shim_numbers_it2,
            datadict_it2,
            shimming_step,
        )
        
button_savefile_dZ = widgets.Button(
    description="Write Z Down to an Excel file.",
    layout=widgets.Layout(width="auto", height="auto"),
)
output_savefile_dZ = widgets.Output()
widget_save_out_dZ = widgets.Output(layout={"border": "1px solid black"})
button_savefile_dZ.on_click(on_button_clicked_savefile_dZ)

box_savefile_dZ = widgets.VBox([button_savefile_dZ, output_savefile_dZ,widget_save_out_dZ])

# %%

first_shim_calculation_menu = [
    module,
    axis,
    fc,
    shim_range_slider,
    stderr_weight,
    button,
    widget_out,
    button_it1,
    widget_stdout_it1,
    box_savefile_it1,
]
second_shim_calculation_menu = [
    module,
    axis,
    fc,
    fc_it2,
    shim_range_slider,
    stderr_weight,
    button,
    widget_out,
    button_it2,
    widget_stdout_it2,
    box_savefile_it2,
    box_savefile_dZ
]